import telebot
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

TOKEN = "8552869258:AAEf39vqNMJeQvV8c3SNQLT1Fgt43AnRneg"

bot = telebot.TeleBot(TOKEN)

@bot.message_handler(commands=['start', 'help'])
def send_welcome(message):
    """Send welcome message"""
    welcome_text = """
👋 Добро пожаловать в Бот Заказов!

📋 Как использовать:

Отправьте данные в этом формате (CSV):
Название,Кол-во,Единица,Кол-во в единице,Цена за штуку

Пример:
Яблоки,10,кг,1,100
Апельсины,5,кг,1,150
Бананы,20,шт,1,50

Я создам красивую таблицу Excel с:
✓ Всеми данными
✓ Общей ценой для каждого продукта
✓ Суммой всех цен снизу

Просто отправь данные! 📊
    """
    bot.reply_to(message, welcome_text)

@bot.message_handler(func=lambda message: True)
def handle_message(message):
    """Receive message and create Excel file with products and prices"""
    try:
        text = message.text
        user_id = message.from_user.id
        
        # Parse data
        lines = text.strip().split('\n')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Заказ"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        total_font = Font(bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add headers
        headers = ["№", "Название", "Кол-во", "Единица", "Кол-во в ед.", "Цена за шт", "Общая цена"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Parse products and add to Excel
        row_idx = 2
        product_rows = []
        
        for line_no, line in enumerate(lines, start=1):
            if not line.strip():
                continue
                
            parts = [p.strip() for p in line.split(',')]
            
            # Parse: Название,Кол-во,Единица,Кол-во_в_единице,Цена
            if len(parts) >= 5:
                try:
                    name = parts[0]
                    quantity = float(parts[1])
                    unit = parts[2]
                    qty_per_unit = float(parts[3])
                    price_per_unit = float(parts[4])
                    
                    # Add to Excel with formatting
                    cell1 = ws.cell(row=row_idx, column=1, value=line_no)
                    cell1.border = border
                    cell1.alignment = Alignment(horizontal='center')
                    
                    cell2 = ws.cell(row=row_idx, column=2, value=name)
                    cell2.border = border
                    cell2.alignment = Alignment(horizontal='left')
                    
                    cell3 = ws.cell(row=row_idx, column=3, value=quantity)
                    cell3.border = border
                    cell3.alignment = Alignment(horizontal='center')
                    cell3.number_format = '0.00'
                    
                    cell4 = ws.cell(row=row_idx, column=4, value=unit)
                    cell4.border = border
                    cell4.alignment = Alignment(horizontal='center')
                    
                    cell5 = ws.cell(row=row_idx, column=5, value=qty_per_unit)
                    cell5.border = border
                    cell5.alignment = Alignment(horizontal='center')
                    cell5.number_format = '0.00'
                    
                    cell6 = ws.cell(row=row_idx, column=6, value=price_per_unit)
                    cell6.border = border
                    cell6.alignment = Alignment(horizontal='right')
                    cell6.number_format = '0.00'
                    
                    # Total price cell - FORMULA (Кол-во × Кол-во в ед. × Цена за шт)
                    total_cell = ws.cell(row=row_idx, column=7, value=f'=C{row_idx}*E{row_idx}*F{row_idx}')
                    total_cell.border = border
                    total_cell.alignment = Alignment(horizontal='right')
                    total_cell.number_format = '0.00'
                    total_cell.font = Font(bold=True)
                    
                    product_rows.append(row_idx)
                    row_idx += 1
                except ValueError:
                    continue
        
        # Add total row
        ws.cell(row=row_idx, column=1, value="").border = border
        
        total_label = ws.cell(row=row_idx, column=6, value="ИТОГО:")
        total_label.fill = total_fill
        total_label.font = total_font
        total_label.alignment = Alignment(horizontal='right')
        total_label.border = border
        
        # Total price cell - FORMULA SUM (сумма всех общих цен)
        if product_rows:
            first_row = product_rows[0]
            last_row = product_rows[-1]
            total_formula = f'=SUM(G{first_row}:G{last_row})'
        else:
            total_formula = '=0'
            
        total_value = ws.cell(row=row_idx, column=7, value=total_formula)
        total_value.fill = total_fill
        total_value.font = total_font
        total_value.border = border
        total_value.number_format = '0.00'
        
        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14
        
        # Set row height for header
        ws.row_dimensions[1].height = 25
        
        # Save file
        filename = f"order_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(os.getcwd(), filename)
        wb.save(filepath)
        
        # Send file to user
        with open(filepath, 'rb') as f:
            bot.send_document(
                message.chat.id, 
                f, 
                caption="✅ Ваш заказ готов!\n\n📊 Таблица содержит формулы для автоматического подсчёта цен"
            )
        
        # Clean up
        os.remove(filepath)
        
    except Exception as e:
        bot.reply_to(message, f"❌ Ошибка: {str(e)}\n\nФормат: Название,Кол-во,Единица,Кол-во в ед,Цена")

print("🤖 Бот запущен...")
print("Отправляйте сообщения боту в Telegram!")
bot.polling()